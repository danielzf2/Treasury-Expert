"""Gera planilha Excel (.xlsx) de simulação NTN-F hedge com DI1.

Modelo 100% auto-contido em FÓRMULAS Excel: basta alterar os inputs
(TIR, quantidade, DUs dos fluxos) e a planilha recalcula tudo.

Abas:
1. "Inputs & Fluxos": inputs editáveis + cashflows com PV, peso, KRD, DV01 em fórmulas
2. "Hedge": 3 modalidades (vencimento, duration, strip) com fórmulas
3. "Cenários": P&L por cenário com fórmulas referenciando DV01 da aba 1

Uso:
    .venv/bin/python scripts/gerar_planilha_hedge.py [--ticker F31] [--qty 2000] [--taxa 13.80]
"""
from __future__ import annotations
import argparse, sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]

# ── Dark navy color palette ─────────────────────────────────────────────────

BG = PatternFill("solid", fgColor="0A0F1C")
CARD = PatternFill("solid", fgColor="111A2B")
INPUT_BG = PatternFill("solid", fgColor="162040")
HDR = PatternFill("solid", fgColor="0D1830")
ORANGE_BG = PatternFill("solid", fgColor="2B1A00")
GREEN_BG = PatternFill("solid", fgColor="0A2015")
RED_BG = PatternFill("solid", fgColor="250A0A")
BLUE_BG = PatternFill("solid", fgColor="0A1530")

W = Font(color="E6EDF3", size=11, name="Calibri")
WB = Font(color="E6EDF3", size=11, name="Calibri", bold=True)
TITLE = Font(color="4C8BF5", size=14, name="Calibri", bold=True)
HDR_F = Font(color="8BA0BE", size=10, name="Calibri", bold=True)
MONO = Font(color="C4D0DC", size=11, name="Consolas")
MONO_B = Font(color="E6EDF3", size=11, name="Consolas", bold=True)
GREEN_F = Font(color="34D399", size=11, name="Consolas", bold=True)
RED_F = Font(color="FF6B6B", size=11, name="Consolas", bold=True)
ORANGE_F = Font(color="F5A623", size=11, name="Consolas", bold=True)
BLUE_F = Font(color="4C8BF5", size=11, name="Calibri", bold=True)
MUTED_F = Font(color="5A6A80", size=10, name="Calibri")
INPUT_F = Font(color="F5A623", size=12, name="Consolas", bold=True)
BORDER = Border(bottom=Side(style="thin", color="1A2540"))


def hdr_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HDR
        cell.font = HDR_F
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def fill_range(ws, r1, c1, r2, c2, fill):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = fill


def col_w(ws, widths):
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ── NTN-F coupon schedule ───────────────────────────────────────────────────

def ntnf_coupon_dus(vcto_year: int, vcto_month: int) -> list[tuple[str, int]]:
    """Gera lista de (label, DU_aproximado) para cupons semestrais de NTN-F.

    NTN-F vence dia 1 do mes. Cupom semestral = 1/Jul e 1/Jan de cada ano.
    DUs sao aproximados (252 DU/ano, 21/mes).
    """
    today = date.today()
    vcto = date(vcto_year, vcto_month, 1)
    # Gera datas de cupom: 1/Jul e 1/Jan antes do vencimento
    coupon_dates = []
    d = vcto
    while True:
        coupon_dates.insert(0, d)
        if d.month == 1:
            d = date(d.year - 1, 7, 1)
        else:
            d = date(d.year, 1, 1)
        if d <= today:
            break

    result = []
    for i, cd in enumerate(coupon_dates):
        delta_days = (cd - today).days
        du = max(1, round(delta_days * 252 / 365))
        label = f"Cupom {i+1}" if i < len(coupon_dates) - 1 else "Principal+Cupom"
        result.append((label, du))
    return result


TICKER_MAP = {
    "F26": (2026, 1), "N26": (2026, 7), "F27": (2027, 1), "N27": (2027, 7),
    "F28": (2028, 1), "N28": (2028, 7), "F29": (2029, 1), "N29": (2029, 7),
    "F30": (2030, 1), "N30": (2030, 7), "F31": (2031, 1), "N31": (2031, 7),
    "F32": (2032, 1), "N32": (2032, 7), "F33": (2033, 1), "F35": (2035, 1),
}


# ── Build aba 1: Inputs & Fluxos ───────────────────────────────────────────

def build_inputs_fluxos(wb, ticker, qty, taxa, flows):
    ws = wb.create_sheet("Inputs & Fluxos")
    ws.sheet_properties.tabColor = "4C8BF5"
    fill_range(ws, 1, 1, 60, 15, BG)

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"NTN-F {ticker} — Fluxos de Caixa, KRD e DV01"
    ws["A1"].font = TITLE

    # === INPUTS (editáveis) ===
    ws["A3"] = "INPUTS (edite aqui)"
    ws["A3"].font = BLUE_F

    labels = [("TIR (% a.a.)", taxa), ("Quantidade", qty), ("Cupom a.a. (%)", 10.0)]
    for i, (lbl, val) in enumerate(labels):
        r = 4 + i
        ws.cell(r, 1, lbl).font = MUTED_F
        ws.cell(r, 2, val).font = INPUT_F
        ws.cell(r, 2).fill = INPUT_BG
        ws.cell(r, 2).number_format = "0.0000" if "TIR" in lbl or "Cupom" in lbl else "#,##0"

    # Cup semestral = (1+cup_aa/100)^0.5 - 1
    ws.cell(7, 1, "Cup Semestral (%)").font = MUTED_F
    ws.cell(7, 2).font = MONO_B
    ws["B7"] = "=((1+B6/100)^0.5-1)*100"
    ws["B7"].number_format = "0.000000"

    # PU total
    ws.cell(8, 1, "PU (R$/un)").font = MUTED_F
    ws.cell(9, 1, "DV01 Total (R$)").font = MUTED_F
    ws.cell(10, 1, "D.Mac (anos)").font = MUTED_F

    # === FLUXOS ===
    flow_start = 12
    headers = ["Fluxo", "DU", "FC (R$/un)", "PV (R$/un)", "Peso (%)", "KRD (anos)",
               "D.Mod Fluxo", "DV01/un (R$)", "DV01 Total (R$)"]
    for i, h in enumerate(headers, 1):
        ws.cell(flow_start, i, h)
    hdr_row(ws, flow_start, len(headers))

    n = len(flows)
    for j, (label, du) in enumerate(flows):
        r = flow_start + 1 + j
        is_last = (j == n - 1)
        ws.cell(r, 1, label).font = W
        ws.cell(r, 2, du).font = MONO
        ws.cell(r, 2).fill = INPUT_BG
        ws.cell(r, 2).alignment = Alignment(horizontal="center")

        # FC: cupom = B7/100*1000; ultimo = (1+B7/100)*1000
        if is_last:
            ws.cell(r, 3).font = MONO
            ws[f"C{r}"] = "=(1+B$7/100)*1000"
        else:
            ws.cell(r, 3).font = MONO
            ws[f"C{r}"] = "=B$7/100*1000"
        ws.cell(r, 3).number_format = "#,##0.00"

        # PV = FC / (1+TIR/100)^(DU/252)
        ws[f"D{r}"] = f"=C{r}/(1+B$4/100)^(B{r}/252)"
        ws.cell(r, 4).font = MONO
        ws.cell(r, 4).number_format = "#,##0.00"

        # Peso = PV / SUM(PV) * 100
        pv_range = f"D${flow_start+1}:D${flow_start+n}"
        ws[f"E{r}"] = f"=D{r}/SUM({pv_range})*100"
        ws.cell(r, 5).font = MONO
        ws.cell(r, 5).number_format = "0.00"

        # KRD = (DU/252) * (Peso/100)
        ws[f"F{r}"] = f"=(B{r}/252)*(E{r}/100)"
        ws.cell(r, 6).font = MONO
        ws.cell(r, 6).number_format = "0.0000"

        # D.Mod fluxo = KRD / (1+TIR/100)
        ws[f"G{r}"] = f"=F{r}/(1+B$4/100)"
        ws.cell(r, 7).font = MONO
        ws.cell(r, 7).number_format = "0.0000"

        # DV01/un = D.Mod * PV * 0.0001
        ws[f"H{r}"] = f"=G{r}*D{r}*0.0001"
        ws.cell(r, 8).font = MONO
        ws.cell(r, 8).number_format = "0.0000"

        # DV01 total = DV01/un * qty
        ws[f"I{r}"] = f"=H{r}*B$5"
        ws.cell(r, 9).font = ORANGE_F
        ws.cell(r, 9).number_format = "#,##0.00"

        for c in range(1, 10):
            ws.cell(r, c).border = BORDER

    # TOTAL row
    tr = flow_start + 1 + n
    ws.cell(tr, 1, "TOTAL").font = WB
    first = flow_start + 1
    last = flow_start + n
    ws[f"D{tr}"] = f"=SUM(D{first}:D{last})"
    ws.cell(tr, 4).font = MONO_B
    ws.cell(tr, 4).number_format = "#,##0.00"
    ws[f"E{tr}"] = f"=SUM(E{first}:E{last})"
    ws.cell(tr, 5).font = MONO_B
    ws[f"F{tr}"] = f"=SUM(F{first}:F{last})"
    ws.cell(tr, 6).font = MONO_B
    ws.cell(tr, 6).number_format = "0.0000"
    ws[f"I{tr}"] = f"=SUM(I{first}:I{last})"
    ws.cell(tr, 9).font = ORANGE_F
    ws.cell(tr, 9).number_format = "#,##0.00"

    # Back-fill PU, DV01 Total, D.Mac
    ws[f"B8"] = f"=D{tr}"
    ws.cell(8, 2).font = MONO_B
    ws.cell(8, 2).number_format = "#,##0.00"
    ws[f"B9"] = f"=I{tr}"
    ws.cell(9, 2).font = ORANGE_F
    ws.cell(9, 2).number_format = "#,##0.00"
    ws[f"B10"] = f"=F{tr}"
    ws.cell(10, 2).font = MONO_B
    ws.cell(10, 2).number_format = "0.0000"

    col_w(ws, {1: 18, 2: 8, 3: 14, 4: 14, 5: 10, 6: 14, 7: 14, 8: 14, 9: 16})
    return ws, flow_start, n


def build_hedge_sheet(wb, n_flows, flow_start):
    ws = wb.create_sheet("Hedge")
    ws.sheet_properties.tabColor = "F5A623"
    fill_range(ws, 1, 1, 50 + n_flows, 10, BG)

    src = "'Inputs & Fluxos'"
    last_flow_row = flow_start + n_flows
    total_row = last_flow_row + 1

    ws.merge_cells("A1:F1")
    ws["A1"] = "Hedge NTN-F com DI1 — 3 Modalidades"
    ws["A1"].font = TITLE

    # Input: DI1 taxa
    ws.cell(3, 1, "DI1 Taxa (% a.a.)").font = MUTED_F
    ws.cell(3, 2, 13.65).font = INPUT_F
    ws.cell(3, 2).fill = INPUT_BG

    # Ref DV01 total NTN-F
    ws.cell(4, 1, "DV01 Total NTN-F").font = MUTED_F
    ws[f"B4"] = f"={src}!B9"
    ws.cell(4, 2).font = ORANGE_F
    ws.cell(4, 2).number_format = "#,##0.00"

    # D.Mac NTN-F
    ws.cell(5, 1, "D.Mac NTN-F (anos)").font = MUTED_F
    ws[f"B5"] = f"={src}!B10"
    ws.cell(5, 2).font = MONO_B
    ws.cell(5, 2).number_format = "0.00"

    # DU vencimento = DU do ultimo fluxo
    ws.cell(6, 1, "DU Vencimento").font = MUTED_F
    ws[f"B6"] = f"={src}!B{last_flow_row}"
    ws.cell(6, 2).font = MONO_B

    # DU Duration = D.Mac * 252
    ws.cell(7, 1, "DU Duration").font = MUTED_F
    ws["B7"] = "=ROUND(B5*252,0)"
    ws.cell(7, 2).font = MONO_B

    # === Modalidades ===
    headers = ["Modalidade", "Vértice DI1 (DU)", "DV01/contrato", "Contratos", "Residual após adotar", "Obs"]
    row = 9
    for i, h in enumerate(headers, 1):
        ws.cell(row, i, h)
    hdr_row(ws, row, len(headers))

    # 1. Vencimento
    r = 10
    ws.cell(r, 1, "1. No Vencimento").font = WB
    ws[f"B{r}"] = "=B6"
    ws.cell(r, 2).font = MONO
    ws[f"C{r}"] = f"=(B{r}/252)/(1+B3/100)*(100000/(1+B3/100)^(B{r}/252))*0.0001"
    ws.cell(r, 3).font = MONO
    ws.cell(r, 3).number_format = "#,##0.00"
    ws[f"D{r}"] = f"=ROUND(B4/C{r},0)"
    ws.cell(r, 4).font = MONO_B
    ws.cell(r, 4).alignment = Alignment(horizontal="center")
    ws[f"E{r}"] = f"=B4-D{r}*C{r}"
    ws.cell(r, 5).font = GREEN_F
    ws.cell(r, 5).number_format = "#,##0.00"
    ws.cell(r, 6, "Aproximação — descasa duration").font = MUTED_F

    # 2. Duration
    r = 11
    ws.cell(r, 1, "2. Na Duration").font = WB
    ws[f"B{r}"] = "=B7"
    ws.cell(r, 2).font = MONO
    ws[f"C{r}"] = f"=(B{r}/252)/(1+B3/100)*(100000/(1+B3/100)^(B{r}/252))*0.0001"
    ws.cell(r, 3).font = MONO
    ws.cell(r, 3).number_format = "#,##0.00"
    ws[f"D{r}"] = f"=ROUND(B4/C{r},0)"
    ws.cell(r, 4).font = MONO_B
    ws.cell(r, 4).alignment = Alignment(horizontal="center")
    ws[f"E{r}"] = f"=B4-D{r}*C{r}"
    ws.cell(r, 5).font = GREEN_F
    ws.cell(r, 5).number_format = "#,##0.00"
    ws.cell(r, 6, "DV01 match no prazo médio — shifts paralelos OK").font = MUTED_F

    # 3. Strip — total de contratos = soma da tabela strip abaixo
    r = 12
    ws.cell(r, 1, "3. Strip (Perfeito)").font = WB
    ws.cell(r, 2, "por fluxo").font = MONO
    first_strip = 16
    last_strip = 15 + n_flows
    ws[f"D{r}"] = f"=SUM(G{first_strip}:G{last_strip})"
    ws.cell(r, 4).font = MONO_B
    ws.cell(r, 4).alignment = Alignment(horizontal="center")
    ws[f"E{r}"] = f"=B4-SUMPRODUCT(F{first_strip}:F{last_strip},G{first_strip}:G{last_strip})"
    ws.cell(r, 5).font = GREEN_F
    ws.cell(r, 5).number_format = "#,##0.00"
    ws.cell(r, 6, "1 DI1 por fluxo — elimina slope/curvatura/convexidade").font = MUTED_F

    for rr in range(10, 13):
        for c in range(1, 7):
            ws.cell(rr, c).border = BORDER

    # === Strip detail ===
    ws.cell(14, 1, "Detalhamento Strip — DI1 por fluxo").font = BLUE_F
    strip_headers = ["Fluxo", "DU", "DV01 Fluxo (R$)", "PU DI1 (R$)", "DV01 DI1/un",
                     "DV01/contrato DI1", "Contratos DI1"]
    sr = 15
    for i, h in enumerate(strip_headers, 1):
        ws.cell(sr, i, h)
    hdr_row(ws, sr, len(strip_headers))

    for j in range(n_flows):
        r = sr + 1 + j
        flow_r = flow_start + 1 + j
        ws[f"A{r}"] = f"={src}!A{flow_r}"
        ws.cell(r, 1).font = W
        ws[f"B{r}"] = f"={src}!B{flow_r}"
        ws.cell(r, 2).font = MONO
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        ws[f"C{r}"] = f"={src}!I{flow_r}"
        ws.cell(r, 3).font = ORANGE_F
        ws.cell(r, 3).number_format = "#,##0.00"
        ws[f"D{r}"] = f"=100000/(1+B$3/100)^(B{r}/252)"
        ws.cell(r, 4).font = MONO
        ws.cell(r, 4).number_format = "#,##0.00"
        ws[f"E{r}"] = f"=(B{r}/252)/(1+B$3/100)*D{r}*0.0001"
        ws.cell(r, 5).font = MONO
        ws.cell(r, 5).number_format = "0.0000"
        ws[f"F{r}"] = f"=E{r}"
        ws.cell(r, 6).font = MONO
        ws.cell(r, 6).number_format = "0.00"
        ws[f"G{r}"] = f"=ROUND(C{r}/F{r},0)"
        ws.cell(r, 7).font = MONO_B
        ws.cell(r, 7).alignment = Alignment(horizontal="center")
        for c in range(1, 8):
            ws.cell(r, c).border = BORDER

    col_w(ws, {1: 22, 2: 14, 3: 16, 4: 14, 5: 14, 6: 18, 7: 16})
    return ws


def build_cenarios_sheet(wb, n_flows, flow_start):
    ws = wb.create_sheet("Cenários P&L")
    ws.sheet_properties.tabColor = "34D399"
    fill_range(ws, 1, 1, 40, 10, BG)

    src = "'Inputs & Fluxos'"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Cenários de Curva — P&L NTN-F vs DI1 no Vencimento"
    ws["A1"].font = TITLE

    ws["A3"] = "Fórmulas de choque (21 DU = curto, 2520 DU = longo):"
    ws["A3"].font = MUTED_F
    ws["A4"] = "Parallel: Δ = magnitude (uniforme em todos os DUs)"
    ws["A4"].font = MUTED_F
    ws["A5"] = "Steepener: Δ = mag × (DU-21)/(2520-21)"
    ws["A5"].font = MUTED_F
    ws["A6"] = "Flattener: Δ = mag × (1 - (DU-21)/(2520-21))"
    ws["A6"].font = MUTED_F

    # P&L NTN-F = sum over flows: -DV01_flow_k × delta_k(DU_k)
    # P&L DI1  = +N_contratos × DV01_di1_per_contrato × delta(DU_vcto)
    # Reference: Hedge!D10 = contratos vcto, Hedge!C10 = DV01/contrato, Hedge!B10 = DU vcto

    headers = ["Cenário", "Mag (bps)", "P&L NTN-F (R$)", "P&L DI1 Vcto (R$)",
               "P&L Total (R$)", "P&L se Strip (≈0)", "Custo Mismatch (R$)", "Risco"]
    row = 8
    for i, h in enumerate(headers, 1):
        ws.cell(row, i, h)
    hdr_row(ws, row, len(headers))

    scenarios = [
        ("Bear Parallel +10", 10, "parallel"),
        ("Bear Parallel +50", 50, "parallel"),
        ("Bull Parallel -50", -50, "parallel"),
        ("Bear Steepener +30", 30, "steep"),
        ("Bear Flattener +30", 30, "flat"),
    ]

    for j, (name, mag, shape) in enumerate(scenarios):
        r = row + 1 + j
        ws.cell(r, 1, name).font = WB
        ws.cell(r, 2, mag).font = MONO
        ws.cell(r, 2).alignment = Alignment(horizontal="center")

        # P&L NTN-F: SUMPRODUCT of (-DV01_flow_k × delta_k)
        # delta_k depends on shape:
        # parallel: mag (constant)
        # steep: mag * (DU-21)/(2520-21)
        # flat: mag * (1 - (DU-21)/(2520-21))
        first_f = flow_start + 1
        last_f = flow_start + n_flows
        du_col = f"{src}!B{first_f}:{src}!B{last_f}"
        dv01_col = f"{src}!I{first_f}:{src}!I{last_f}"

        if shape == "parallel":
            # All flows get same delta = mag
            ws[f"C{r}"] = f"=-SUMPRODUCT({dv01_col})*B{r}"
        elif shape == "steep":
            # delta_k = mag * (DU_k - 21) / (2520 - 21)
            ws[f"C{r}"] = f"=-SUMPRODUCT({dv01_col},({du_col}-21)/(2520-21))*B{r}"
        elif shape == "flat":
            # delta_k = mag * (1 - (DU_k - 21) / (2520 - 21))
            ws[f"C{r}"] = f"=-SUMPRODUCT({dv01_col},(1-({du_col}-21)/(2520-21)))*B{r}"
        ws.cell(r, 3).font = MONO
        ws.cell(r, 3).number_format = "#,##0"

        # P&L DI1: +N_contratos × DV01/contrato × delta_at_vcto
        # Ref: Hedge!D10 = N, Hedge!C10 = DV01/ct, Hedge!B10 = DU vcto
        if shape == "parallel":
            ws[f"D{r}"] = f"='Hedge'!D10*'Hedge'!C10*B{r}"
        elif shape == "steep":
            ws[f"D{r}"] = f"='Hedge'!D10*'Hedge'!C10*B{r}*('Hedge'!B10-21)/(2520-21)"
        elif shape == "flat":
            ws[f"D{r}"] = f"='Hedge'!D10*'Hedge'!C10*B{r}*(1-('Hedge'!B10-21)/(2520-21))"
        ws.cell(r, 4).font = MONO
        ws.cell(r, 4).number_format = "#,##0"

        # Total
        ws[f"E{r}"] = f"=C{r}+D{r}"
        ws.cell(r, 5).font = MONO_B
        ws.cell(r, 5).number_format = "#,##0"

        # Strip ≈ 0
        ws.cell(r, 6, 0).font = GREEN_F

        # Custo mismatch = Total - Strip
        ws[f"G{r}"] = f"=E{r}-F{r}"
        ws.cell(r, 7).font = ORANGE_F
        ws.cell(r, 7).number_format = "#,##0"

        # Risco label
        is_slope = shape in ("steep", "flat")
        ws.cell(r, 8, "RISCO SLOPE" if is_slope else "OK (DV01 match)").font = RED_F if is_slope else GREEN_F

        for c in range(1, 9):
            ws.cell(r, c).border = BORDER

    col_w(ws, {1: 24, 2: 12, 3: 16, 4: 18, 5: 16, 6: 16, 7: 18, 8: 20})
    return ws


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", default="F31")
    ap.add_argument("--qty", type=int, default=2000)
    ap.add_argument("--taxa", type=float, default=13.80)
    ap.add_argument("--output", default=None)
    args = ap.parse_args()

    t = TICKER_MAP.get(args.ticker)
    if not t:
        print(f"Ticker {args.ticker} não reconhecido. Use: {', '.join(TICKER_MAP.keys())}")
        sys.exit(1)

    flows = ntnf_coupon_dus(t[0], t[1])
    print(f"NTN-F {args.ticker}: {len(flows)} fluxos, vcto {t[0]}-{t[1]:02d}-01")

    wb = Workbook()
    wb.remove(wb.active)

    ws1, flow_start, n = build_inputs_fluxos(wb, args.ticker, args.qty, args.taxa, flows)
    build_hedge_sheet(wb, n, flow_start)
    build_cenarios_sheet(wb, n, flow_start)

    out = args.output or str(ROOT / "planilhas" / f"NTN-F_{args.ticker}_hedge_{args.qty}.xlsx")
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Planilha salva: {out}")
    print("Tudo em FÓRMULAS — edite TIR/Qty/DUs e recalcula automaticamente.")


if __name__ == "__main__":
    main()
